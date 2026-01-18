"""
Script để xuất dữ liệu ECG từ API ra file Excel với hình ảnh ECG.
"""

import json
import io
import os
from datetime import datetime
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')  # Use non-interactive backend
import numpy as np
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# ============== CẤU HÌNH ==============
CASES_FOLDER = "36_cases_rag"  # Folder chứa các file case_*.json


def load_cases_from_folder(folder_path):
    """Load tất cả các file case từ folder 36_cases."""
    if not os.path.exists(folder_path):
        print(f"Folder '{folder_path}' không tồn tại!")
        return []
    
    case_files = sorted([f for f in os.listdir(folder_path) if f.startswith('case_') and f.endswith('.json')])
    
    cases = []
    for case_file in case_files:
        file_path = os.path.join(folder_path, case_file)
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                case_data = json.load(f)
                cases.append(case_data)
        except Exception as e:
            print(f"Lỗi khi đọc file {case_file}: {e}")
    
    print(f"Đã load {len(cases)} case(s) từ folder '{folder_path}'")
    return cases


def create_ecg_plot(signal, title, sampling_rate=130, figsize=(8, 3)):
    """Tạo hình ảnh ECG từ signal."""
    fig, ax = plt.subplots(figsize=figsize)
    
    # Tạo trục thời gian
    duration = len(signal) / sampling_rate
    time_axis = np.linspace(0, duration, len(signal))
    
    # Vẽ ECG
    ax.plot(time_axis, signal, 'b-', linewidth=0.8)
    ax.set_xlabel('Thời gian (giây)', fontsize=10)
    ax.set_ylabel('Biên độ (mV)', fontsize=10)
    ax.set_title(title, fontsize=11, fontweight='bold')
    ax.grid(True, alpha=0.3)
    ax.set_xlim(0, duration)
    
    # Thêm nền giấy ECG
    ax.set_facecolor('#fff5f5')
    
    plt.tight_layout()
    
    # Lưu vào buffer
    img_buffer = io.BytesIO()
    fig.savefig(img_buffer, format='png', dpi=100, bbox_inches='tight', 
                facecolor='white', edgecolor='none')
    img_buffer.seek(0)
    plt.close(fig)
    
    return img_buffer


def format_explanation(explanation_obj):
    """Format explanation object thành text đẹp cho Excel."""
    if not explanation_obj or 'explanation' not in explanation_obj:
        return "Không có dữ liệu"
    
    exp = explanation_obj.get('explanation', {})
    
    # Format các phần của explanation
    parts = []
    
    # Tóm tắt
    if exp.get('summary'):
        parts.append(f"📋 TÓM TẮT:\n{exp['summary']}")
    
    # Chi tiết
    if exp.get('details'):
        parts.append(f"\n📝 CHI TIẾT:\n{exp['details']}")
    
    # Khuyến nghị
    if exp.get('recommendations'):
        parts.append(f"\n💊 KHUYẾN NGHỊ:\n{exp['recommendations']}")
    
    # Mức độ rủi ro
    if exp.get('risk_level'):
        risk_map = {
            'low': '🟢 Thấp',
            'medium': '🟡 Trung bình', 
            'high': '🔴 Cao'
        }
        risk = risk_map.get(exp['risk_level'], exp['risk_level'])
        parts.append(f"\n⚠️ MỨC ĐỘ RỦI RO: {risk}")
    
    # Bước tiếp theo
    if exp.get('next_steps'):
        parts.append(f"\n👉 BƯỚC TIẾP THEO:\n{exp['next_steps']}")
    
    return '\n'.join(parts) if parts else "Không có dữ liệu"


def create_excel_report(data_list, output_file):
    """Tạo file Excel với dữ liệu ECG."""
    # Sắp xếp data_list theo loại bệnh để các case cùng bệnh nằm kế nhau
    sorted_data = sorted(data_list, key=lambda x: (
        x.get('prediction', {}).get('diagnosis', 'ZZZ'),  # ZZZ để đẩy các mục không có diagnosis xuống cuối
        x.get('id', '')
    ))
    
    wb = Workbook()
    ws = wb.active
    ws.title = "ECG Report"
    
    # Định nghĩa styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Tạo header
    headers = ["STT", "Hình ảnh ECG Raw", "Hình ảnh ECG Denoised", "Loại bệnh lý", "Đánh giá từ LLM"]
    
    # Đặt độ rộng cột
    column_widths = [8, 45, 45, 30, 80]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Viết header
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    ws.row_dimensions[1].height = 30
    
    # Tạo thư mục tạm để lưu hình ảnh
    temp_img_dir = "temp_ecg_images"
    os.makedirs(temp_img_dir, exist_ok=True)
    
    # Viết dữ liệu (sử dụng sorted_data thay vì data_list)
    for idx, data in enumerate(sorted_data, 1):
        row = idx + 1
        
        # Đặt chiều cao hàng để chứa hình ảnh
        ws.row_dimensions[row].height = 180
        
        # STT
        cell = ws.cell(row=row, column=1, value=idx)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        
        # Hình ảnh ECG Raw
        ecg_recording = data.get('ecgRecording', data.get('ecg_recording'))
        if ecg_recording and ecg_recording.get('rawData'):
            raw_signal = ecg_recording['rawData'].get('signal', [])
            if raw_signal:
                sampling_rate = ecg_recording.get('samplingRate', 130)
                img_buffer = create_ecg_plot(raw_signal, "ECG Raw Signal", sampling_rate)
                
                # Lưu tạm và chèn vào Excel
                img_path = os.path.join(temp_img_dir, f"raw_{idx}.png")
                with open(img_path, 'wb') as f:
                    f.write(img_buffer.getvalue())
                
                img = XLImage(img_path)
                img.width = 300
                img.height = 120
                ws.add_image(img, f'B{row}')
        
        ws.cell(row=row, column=2).border = thin_border
        
        # Hình ảnh ECG Denoised
        if ecg_recording and ecg_recording.get('denoisedData'):
            denoised_signal = ecg_recording['denoisedData'].get('signal', [])
            if denoised_signal:
                sampling_rate = ecg_recording.get('samplingRate', 130)
                img_buffer = create_ecg_plot(denoised_signal, "ECG Denoised Signal", sampling_rate)
                
                # Lưu tạm và chèn vào Excel
                img_path = os.path.join(temp_img_dir, f"denoised_{idx}.png")
                with open(img_path, 'wb') as f:
                    f.write(img_buffer.getvalue())
                
                img = XLImage(img_path)
                img.width = 300
                img.height = 120
                ws.add_image(img, f'C{row}')
        
        ws.cell(row=row, column=3).border = thin_border
        
        # Loại bệnh lý
        diagnosis = "Không có dữ liệu"
        if data.get('prediction') and data['prediction'].get('diagnosis'):
            diagnosis = data['prediction']['diagnosis']
            probability = data['prediction'].get('probability', 0)
            diagnosis = f"{diagnosis}\n\n(Độ tin cậy: {probability*100:.2f}%)"
        
        cell = ws.cell(row=row, column=4, value=diagnosis)
        cell.alignment = cell_alignment
        cell.border = thin_border
        
        # Đánh giá từ LLM
        explanation_text = format_explanation(data.get('explanation'))
        cell = ws.cell(row=row, column=5, value=explanation_text)
        cell.alignment = cell_alignment
        cell.border = thin_border
    
    # Lưu file Excel
    wb.save(output_file)
    print(f"Đã lưu file Excel: {output_file}")
    
    # Xóa thư mục tạm
    import shutil
    try:
        shutil.rmtree(temp_img_dir)
        print("Đã xóa thư mục tạm chứa hình ảnh")
    except Exception as e:
        print(f"Không thể xóa thư mục tạm: {e}")


def main():
    print("=" * 60)
    print("SCRIPT XUẤT DỮ LIỆU ECG RA FILE EXCEL")
    print("=" * 60)
    
    # Bước 1: Load tất cả cases từ folder 36_cases
    print(f"\n[1/2] Đang đọc dữ liệu từ folder '{CASES_FOLDER}'...")
    cases = load_cases_from_folder(CASES_FOLDER)
    
    if not cases:
        print(f"Không tìm thấy case nào trong folder '{CASES_FOLDER}'. Kết thúc.")
        return
    
    # Thống kê các loại bệnh
    disease_count = {}
    for case in cases:
        diagnosis = case.get('prediction', {}).get('diagnosis', 'Không xác định')
        disease_count[diagnosis] = disease_count.get(diagnosis, 0) + 1
    
    print("\nThống kê theo loại bệnh:")
    for disease, count in sorted(disease_count.items()):
        print(f"  - {disease}: {count} case(s)")
    
    # Bước 2: Tạo file Excel
    print(f"\n[2/2] Đang tạo file Excel từ {len(cases)} case(s)...")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f"ecg_report_{timestamp}.xlsx"
    create_excel_report(cases, output_file)
    
    print("\n" + "=" * 60)
    print("HOÀN THÀNH!")
    print(f"File Excel đã được lưu: {output_file}")
    print("=" * 60)


if __name__ == "__main__":
    main()
